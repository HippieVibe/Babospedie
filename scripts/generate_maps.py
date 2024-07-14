import csv
import os
import re
import warnings
from dataclasses import dataclass
from datetime import date
from io import BytesIO, TextIOWrapper
from pathlib import Path
from statistics import mean
from tempfile import NamedTemporaryFile, TemporaryFile
from time import sleep
from typing import Any
from urllib.parse import urlsplit, urlunsplit
from zipfile import ZipFile

import requests
import requests_cache
from jenkspy import jenks_breaks
from openpyxl import load_workbook
from pygal.config import Config
from pygal.style import Style
from pygal_maps_fr.maps import DEPARTMENTS
from pygal_maps_fr.maps import Departments as FrenchMapDepartments

THIS_SCRIPT_LOCATION = Path(os.path.realpath(__file__)).parent
HTTP_TIMEOUT = 10

requests_cache.install_cache("generate_maps")


class CityError(Exception):
    """Raised when something went wrong with the geocoding API or its computation."""


class ClimateError(Exception):
    """Raised when something went wrong with the climate API or its computation."""


class AirQualityError(Exception):
    """Raised when something went wrong with the air quality API or its computation."""


class NaturalDisasterError(Exception):
    """Raised when something went wrong with the processing of the GASPAR database."""


class SoilPollutionError(Exception):
    """Raised when something went wrong with the processing of the BASOL database."""


@dataclass(frozen=True)
class City:
    name: str
    latitude: float
    longitude: float
    elevation: int
    population: int
    departement: str


@dataclass(frozen=True)
class AirQuality:
    european_aqi_min: int
    european_aqi_pm2_5_min: int
    european_aqi_pm10_min: int
    european_aqi_nitrogen_dioxide_min: int
    european_aqi_ozone_min: int
    european_aqi_sulphur_dioxide_min: int

    european_aqi_max: int
    european_aqi_pm2_5_max: int
    european_aqi_pm10_max: int
    european_aqi_nitrogen_dioxide_max: int
    european_aqi_ozone_max: int
    european_aqi_sulphur_dioxide_max: int

    european_aqi_mean: float
    european_aqi_pm2_5_mean: float
    european_aqi_pm10_mean: float
    european_aqi_nitrogen_dioxide_mean: float
    european_aqi_ozone_mean: float
    european_aqi_sulphur_dioxide_mean: float


@dataclass(frozen=True)
class Weather:
    # Maximum air temperature at 2 meters above ground.
    temperature_2m_max: float

    # Minimum air temperature at 2 meters above ground.
    temperature_2m_min: float

    # Mean temperature during the day, 2 meters above the ground.
    temperature_2m_mean: float

    # Maximum apparent air temperature.
    apparent_temperature_max: float

    # Minimum apparent air temperature.
    apparent_temperature_min: float

    # Mean apparent air temperature.
    apparent_temperature_mean: float

    # Maximum wind speed on a day.
    wind_speed_10m_max: float

    # Number of seconds of daylight per day.
    daylight_duration: float

    # The number of seconds of sunshine per day is determined by calculating direct
    # normalized irradiance exceeding 120 W/m², following the WMO definition.
    # Sunshine duration will consistently be less than daylight duration due to dawn and
    # dusk.
    sunshine_duration: float

    # Sum of daily rain.
    rainfall_sum: float

    # Sum of daily snowfall.
    snowfall_sum: float

    # The number of hours with rain.
    precipitation_hours: float

    # Evapotranspiration of a well watered grass field.
    # Commonly used to estimate the required irrigation for plants.
    et0_fao_evapotranspiration: float


def requests_get_meteo(url: str, params: dict) -> dict[str, Any]:
    # Free trial is available to get an API key with more requests per day.
    if "METEO_API_KEY" in os.environ:
        params["apikey"] = os.environ["METEO_API_KEY"]
        # Because the subdomain is not the same when using an API key.
        tokens = list(urlsplit(url))
        tokens[1] = "customer-" + tokens[1]
        url = urlunsplit(tokens)

    resp = requests.get(url, params=params, timeout=HTTP_TIMEOUT)
    json_content = resp.json()

    if resp.status_code != 200:
        if "Minutely API request limit exceeded" in json_content["reason"]:
            print("Minutely API request limit reached, waiting for 60 seconds...")
            sleep(60)
            return requests_get_meteo(url, params)

        if "Hourly API request limit exceeded" in json_content["reason"]:
            print("Hourly API request limit reached, waiting for 1 hour...")
            sleep(3600)
            return requests_get_meteo(url, params)

        if "Daily API request limit exceeded" in json_content["reason"]:
            print("Daily API request limit reached, waiting for 1 day...")
            sleep(3600 * 24)
            return requests_get_meteo(url, params)

        resp.raise_for_status()

    return resp.json()


def get_weather(city: City, start_date: date, end_date: date) -> dict[date, Weather]:
    json_content = requests_get_meteo(
        "https://archive-api.open-meteo.com/v1/archive",
        params={
            "latitude": city.latitude,
            "longitude": city.longitude,
            "start_date": str(start_date),
            "end_date": str(end_date),
            "daily": [
                "temperature_2m_max",
                "temperature_2m_min",
                "temperature_2m_mean",
                "apparent_temperature_max",
                "apparent_temperature_min",
                "apparent_temperature_mean",
                "daylight_duration",
                "sunshine_duration",
                "wind_speed_10m_max",
                "rain_sum",
                "snowfall_sum",
                "precipitation_hours",
                "et0_fao_evapotranspiration",
            ],
        },
    )

    json_content = json_content["daily"]

    time = json_content["time"]
    temperature_2m_max = json_content["temperature_2m_max"]
    temperature_2m_min = json_content["temperature_2m_min"]
    temperature_2m_mean = json_content["temperature_2m_mean"]
    apparent_temperature_max = json_content["apparent_temperature_max"]
    apparent_temperature_min = json_content["apparent_temperature_min"]
    apparent_temperature_mean = json_content["apparent_temperature_mean"]
    daylight_duration = json_content["daylight_duration"]
    sunshine_duration = json_content["sunshine_duration"]
    wind_speed_10m_max = json_content["wind_speed_10m_max"]
    rain_sum = json_content["rain_sum"]
    snowfall_sum = json_content["snowfall_sum"]
    precipitation_hours = json_content["precipitation_hours"]
    et0_fao_evapotranspiration = json_content["et0_fao_evapotranspiration"]

    measurements: dict[date, Weather] = {}
    for i in range(len(time)):
        year, month, day = time[i].split("-")
        year, month, day = int(year), int(month), int(day)

        # Sometimes there's a hole in the data (sensor failure?)
        # it is quite rare but if it happens, we just throw the data away
        if temperature_2m_max[i] is None:
            print(f"WARNING: temperature_2m_max is None, skipping {time[i]}")
            continue

        if temperature_2m_min[i] is None:
            print(f"WARNING: temperature_2m_min is None, skipping {time[i]}")
            continue

        if temperature_2m_mean[i] is None:
            print(f"WARNING: temperature_2m_mean is None, skipping {time[i]}")
            continue

        if apparent_temperature_max[i] is None:
            print(f"WARNING: apparent_temperature_max is None, skipping {time[i]}")
            continue

        if apparent_temperature_min[i] is None:
            print(f"WARNING: apparent_temperature_min is None, skipping {time[i]}")
            continue

        if apparent_temperature_mean[i] is None:
            print(f"WARNING: apparent_temperature_mean is None, skipping {time[i]}")
            continue

        if wind_speed_10m_max[i] is None:
            print(f"WARNING: wind_speed_10m_max is None, skipping {time[i]}")
            continue

        if daylight_duration[i] is None:
            print(f"WARNING: daylight_duration is None, skipping {time[i]}")
            continue

        if sunshine_duration[i] is None:
            print(f"WARNING: sunshine_duration is None, skipping {time[i]}")
            continue

        if rain_sum[i] is None:
            print(f"WARNING: rain_sum is None, skipping {time[i]}")
            continue

        if snowfall_sum[i] is None:
            print(f"WARNING: snowfall_sum is None, skipping {time[i]}")
            continue

        if precipitation_hours[i] is None:
            print(f"WARNING: precipitation_hours is None, skipping {time[i]}")
            continue

        if et0_fao_evapotranspiration[i] is None:
            print(f"WARNING: et0_fao_evapotranspiration is None, skipping {time[i]}")
            continue

        dw = Weather(
            temperature_2m_max[i],
            temperature_2m_min[i],
            temperature_2m_mean[i],
            apparent_temperature_max[i],
            apparent_temperature_min[i],
            apparent_temperature_mean[i],
            wind_speed_10m_max[i],
            daylight_duration[i] / 3600,  # convert to hours
            sunshine_duration[i] / 3600,  # convert to hours
            rain_sum[i],
            snowfall_sum[i] * 10,  # convert to mm
            precipitation_hours[i],
            et0_fao_evapotranspiration[i],
        )
        measurements[date(year, month, day)] = dw

    return measurements


def compute_average_weather(weather_measurements: list[Weather]) -> Weather:
    temperature_2m_max_values = [wm.temperature_2m_max for wm in weather_measurements]
    temperature_2m_min_values = [wm.temperature_2m_min for wm in weather_measurements]
    temperature_2m_mean_values = [wm.temperature_2m_mean for wm in weather_measurements]
    apparent_temperature_max_values = [
        wm.apparent_temperature_max for wm in weather_measurements
    ]
    apparent_temperature_min_values = [
        wm.apparent_temperature_min for wm in weather_measurements
    ]
    apparent_temperature_mean_values = [
        wm.apparent_temperature_mean for wm in weather_measurements
    ]
    wind_speed_10m_max_values = [wm.wind_speed_10m_max for wm in weather_measurements]
    daylight_duration_values = [wm.daylight_duration for wm in weather_measurements]
    sunshine_duration_values = [wm.sunshine_duration for wm in weather_measurements]
    rain_sum_values = [wm.rainfall_sum for wm in weather_measurements]
    snowfall_sum_values = [wm.snowfall_sum for wm in weather_measurements]
    precipitation_hours_values = [wm.precipitation_hours for wm in weather_measurements]
    et0_fao_evapotranspiration_values = [
        wm.et0_fao_evapotranspiration for wm in weather_measurements
    ]

    return Weather(
        temperature_2m_max=mean(temperature_2m_max_values),
        temperature_2m_min=mean(temperature_2m_min_values),
        temperature_2m_mean=mean(temperature_2m_mean_values),
        apparent_temperature_max=mean(apparent_temperature_max_values),
        apparent_temperature_min=mean(apparent_temperature_min_values),
        apparent_temperature_mean=mean(apparent_temperature_mean_values),
        wind_speed_10m_max=mean(wind_speed_10m_max_values),
        daylight_duration=mean(daylight_duration_values),
        sunshine_duration=mean(sunshine_duration_values),
        rainfall_sum=mean(rain_sum_values),
        snowfall_sum=mean(snowfall_sum_values),
        precipitation_hours=mean(precipitation_hours_values),
        et0_fao_evapotranspiration=mean(et0_fao_evapotranspiration_values),
    )


def compute_average_season_weather(
    weather_measurements: dict[date, Weather],
) -> tuple[Weather, Weather]:
    """Process winter and summer means given all the input measurements.

    Winter is considered to last over december, january and february.
    Summer is considered to last over june, july and august.
    """
    years_to_process = {date.year for date in weather_measurements}

    winter_temperature_2m_max_values = []
    winter_temperature_2m_min_values = []
    winter_temperature_2m_values = []
    winter_apparent_temparature_max_values = []
    winter_apparent_temparature_min_values = []
    winter_apparent_temparature_mean_values = []
    winter_wind_speed_10m_max_values = []
    winter_daylight_duration_values = []
    winter_sunshine_duration_values = []
    winter_rain_sum_values = []
    winter_snowfall_sum_values = []
    winter_precipitation_hours_values = []
    winter_et0_fao_evapotranspiration_values = []

    summer_temperature_2m_max_values = []
    summer_temperature_2m_min_values = []
    summer_temperature_2m_values = []
    summer_apparent_temparature_max_values = []
    summer_apparent_temparature_min_values = []
    summer_apparent_temparature_mean_values = []
    summer_wind_speed_10m_max_values = []
    summer_daylight_duration_values = []
    summer_sunshine_duration_values = []
    summer_rain_sum_values = []
    summer_snowfall_sum_values = []
    summer_precipitation_hours_values = []
    summer_et0_fao_evapotranspiration_values = []

    for year in years_to_process:
        weather_measurements_for_the_given_year = {
            date: wm for date, wm in weather_measurements.items() if date.year == year
        }
        weather_measurements_in_december = [
            wm
            for date, wm in weather_measurements_for_the_given_year.items()
            if date.month == 12
        ]
        weather_measurements_in_january = [
            wm
            for date, wm in weather_measurements_for_the_given_year.items()
            if date.month == 1
        ]
        weather_measurements_in_february = [
            wm
            for date, wm in weather_measurements_for_the_given_year.items()
            if date.month == 2
        ]
        weather_measurements_in_june = [
            wm
            for date, wm in weather_measurements_for_the_given_year.items()
            if date.month == 6
        ]
        weather_measurements_in_july = [
            wm
            for date, wm in weather_measurements_for_the_given_year.items()
            if date.month == 7
        ]
        weather_measurements_in_august = [
            wm
            for date, wm in weather_measurements_for_the_given_year.items()
            if date.month == 8
        ]

        winter_measurements_to_process = (
            weather_measurements_in_december
            + weather_measurements_in_january
            + weather_measurements_in_february
        )

        summer_measurements_to_process = (
            weather_measurements_in_june
            + weather_measurements_in_july
            + weather_measurements_in_august
        )

        # Rainfall and snowfall are processed separately because we want
        # the total fall over the season instead of performing a daily mean.
        winter_rain_sum_values.append(
            sum(wm.rainfall_sum for wm in winter_measurements_to_process),
        )
        winter_snowfall_sum_values.append(
            sum(wm.snowfall_sum for wm in winter_measurements_to_process),
        )
        winter_et0_fao_evapotranspiration_values.append(
            sum(wm.et0_fao_evapotranspiration for wm in winter_measurements_to_process),
        )
        summer_rain_sum_values.append(
            sum(wm.rainfall_sum for wm in summer_measurements_to_process),
        )
        summer_snowfall_sum_values.append(
            sum(wm.snowfall_sum for wm in summer_measurements_to_process),
        )
        summer_et0_fao_evapotranspiration_values.append(
            sum(wm.et0_fao_evapotranspiration for wm in summer_measurements_to_process),
        )

        for weather_measurement in winter_measurements_to_process:
            winter_temperature_2m_max_values.append(
                weather_measurement.temperature_2m_max,
            )
            winter_temperature_2m_min_values.append(
                weather_measurement.temperature_2m_min,
            )
            winter_temperature_2m_values.append(weather_measurement.temperature_2m_mean)
            winter_apparent_temparature_max_values.append(
                weather_measurement.apparent_temperature_max,
            )
            winter_apparent_temparature_min_values.append(
                weather_measurement.apparent_temperature_min,
            )
            winter_apparent_temparature_mean_values.append(
                weather_measurement.apparent_temperature_mean,
            )
            winter_wind_speed_10m_max_values.append(
                weather_measurement.wind_speed_10m_max,
            )
            winter_daylight_duration_values.append(
                weather_measurement.daylight_duration,
            )
            winter_sunshine_duration_values.append(
                weather_measurement.sunshine_duration,
            )
            winter_precipitation_hours_values.append(
                weather_measurement.precipitation_hours,
            )

        for weather_measurement in summer_measurements_to_process:
            summer_temperature_2m_max_values.append(
                weather_measurement.temperature_2m_max,
            )
            summer_temperature_2m_min_values.append(
                weather_measurement.temperature_2m_min,
            )
            summer_temperature_2m_values.append(weather_measurement.temperature_2m_mean)
            summer_apparent_temparature_max_values.append(
                weather_measurement.apparent_temperature_max,
            )
            summer_apparent_temparature_min_values.append(
                weather_measurement.apparent_temperature_min,
            )
            summer_apparent_temparature_mean_values.append(
                weather_measurement.apparent_temperature_mean,
            )
            summer_wind_speed_10m_max_values.append(
                weather_measurement.wind_speed_10m_max,
            )
            summer_daylight_duration_values.append(
                weather_measurement.daylight_duration,
            )
            summer_sunshine_duration_values.append(
                weather_measurement.sunshine_duration,
            )
            summer_precipitation_hours_values.append(
                weather_measurement.precipitation_hours,
            )

    return (
        Weather(
            temperature_2m_max=mean(winter_temperature_2m_max_values),
            temperature_2m_min=mean(winter_temperature_2m_min_values),
            temperature_2m_mean=mean(winter_temperature_2m_values),
            apparent_temperature_max=mean(winter_apparent_temparature_max_values),
            apparent_temperature_min=mean(winter_apparent_temparature_min_values),
            apparent_temperature_mean=mean(winter_apparent_temparature_mean_values),
            wind_speed_10m_max=mean(winter_wind_speed_10m_max_values),
            daylight_duration=mean(winter_daylight_duration_values),
            sunshine_duration=mean(winter_sunshine_duration_values),
            rainfall_sum=mean(winter_rain_sum_values),
            snowfall_sum=mean(winter_snowfall_sum_values),
            precipitation_hours=mean(winter_precipitation_hours_values),
            et0_fao_evapotranspiration=mean(winter_et0_fao_evapotranspiration_values),
        ),
        Weather(
            temperature_2m_max=mean(summer_temperature_2m_max_values),
            temperature_2m_min=mean(summer_temperature_2m_min_values),
            temperature_2m_mean=mean(summer_temperature_2m_values),
            apparent_temperature_max=mean(summer_apparent_temparature_max_values),
            apparent_temperature_min=mean(summer_apparent_temparature_min_values),
            apparent_temperature_mean=mean(summer_apparent_temparature_mean_values),
            wind_speed_10m_max=mean(summer_wind_speed_10m_max_values),
            daylight_duration=mean(summer_daylight_duration_values),
            sunshine_duration=mean(summer_sunshine_duration_values),
            rainfall_sum=mean(summer_rain_sum_values),
            snowfall_sum=mean(summer_snowfall_sum_values),
            precipitation_hours=mean(summer_precipitation_hours_values),
            et0_fao_evapotranspiration=mean(summer_et0_fao_evapotranspiration_values),
        ),
    )


def get_air_quality_mean(city: City, start_date: date, end_date: date) -> AirQuality:
    json_content = requests_get_meteo(
        "https://air-quality-api.open-meteo.com/v1/air-quality",
        params={
            "latitude": city.latitude,
            "longitude": city.longitude,
            "hourly": [
                "european_aqi,european_aqi_pm2_5",
                "european_aqi_pm10",
                "european_aqi_nitrogen_dioxide",
                "european_aqi_ozone",
                "european_aqi_sulphur_dioxide",
            ],
            "start_date": str(start_date),
            "end_date": str(end_date),
        },
    )

    json_content = json_content["hourly"]

    european_aqi = [v for v in json_content["european_aqi"] if v is not None]
    european_aqi_pm2_5 = [
        v for v in json_content["european_aqi_pm2_5"] if v is not None
    ]
    european_aqi_pm10 = [v for v in json_content["european_aqi_pm10"] if v is not None]
    european_aqi_nitrogen_dioxide = [
        v for v in json_content["european_aqi_nitrogen_dioxide"] if v is not None
    ]
    european_aqi_ozone = [
        v for v in json_content["european_aqi_ozone"] if v is not None
    ]
    european_aqi_sulphur_dioxide = [
        v for v in json_content["european_aqi_sulphur_dioxide"] if v is not None
    ]

    return AirQuality(
        european_aqi_min=min(european_aqi),
        european_aqi_pm2_5_min=min(european_aqi_pm2_5),
        european_aqi_pm10_min=min(european_aqi_pm10),
        european_aqi_nitrogen_dioxide_min=min(european_aqi_nitrogen_dioxide),
        european_aqi_ozone_min=min(european_aqi_ozone),
        european_aqi_sulphur_dioxide_min=min(european_aqi_sulphur_dioxide),
        european_aqi_max=max(european_aqi),
        european_aqi_pm2_5_max=max(european_aqi_pm2_5),
        european_aqi_pm10_max=max(european_aqi_pm10),
        european_aqi_nitrogen_dioxide_max=max(european_aqi_nitrogen_dioxide),
        european_aqi_ozone_max=max(european_aqi_ozone),
        european_aqi_sulphur_dioxide_max=max(european_aqi_sulphur_dioxide),
        european_aqi_mean=mean(european_aqi),
        european_aqi_pm2_5_mean=mean(european_aqi_pm2_5),
        european_aqi_pm10_mean=mean(european_aqi_pm10),
        european_aqi_nitrogen_dioxide_mean=mean(european_aqi_nitrogen_dioxide),
        european_aqi_ozone_mean=mean(european_aqi_ozone),
        european_aqi_sulphur_dioxide_mean=mean(european_aqi_sulphur_dioxide),
    )


def get_referenced_natural_disaster_count() -> dict[str, int]:
    resp = requests.get(
        "http://files.georisques.fr/GASPAR/gaspar.zip",
        timeout=HTTP_TIMEOUT,
    )
    f = ZipFile(BytesIO(resp.content))

    natural_disasters_count_per_department = {code: 0 for code in DEPARTMENTS}

    with f.open("catnat_gaspar.csv", "r") as csv_file:
        csv_reader = csv.reader(TextIOWrapper(csv_file), delimiter=";")
        next(csv_reader)  # skip header

        for row in csv_reader:
            city_code = row[1]
            department_code = (
                city_code[:3] if city_code.startswith("97") else city_code[:2]
            )

            if department_code in natural_disasters_count_per_department:
                natural_disasters_count_per_department[department_code] += 1
            else:
                natural_disasters_count_per_department[department_code] = 1

    return natural_disasters_count_per_department


def get_soil_pollution_incidents_count() -> dict[str, int]:
    """Return how many pollution incidents happened in the given city.

    https://www.data.gouv.fr/en/datasets/base-des-sols-pollues/
    """
    incidents_count_per_department = {code: 0 for code in DEPARTMENTS}

    resp = requests.get(
        "https://www.georisques.gouv.fr/webappReport/ws/infosols/export/excel?national=true",
        timeout=HTTP_TIMEOUT,
    )
    xlsx_file = BytesIO(resp.content)
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            category=UserWarning,
            module=re.escape("openpyxl.styles.stylesheet"),
        )
        workbook = load_workbook(xlsx_file)

    worksheet = workbook.active

    if worksheet is None:
        error = "Unable to find default worksheet"
        raise SoilPollutionError(error)

    for row in worksheet.iter_rows(min_row=2):
        insee_code = row[8].value

        if not isinstance(insee_code, str):
            error = f"Unexpected cell value: {insee_code}"
            raise SoilPollutionError(error)

        department_code = (
            insee_code[:3] if insee_code.startswith("97") else insee_code[:2]
        )

        if department_code in incidents_count_per_department:
            incidents_count_per_department[department_code] += 1
        else:
            incidents_count_per_department[department_code] = 1

    return incidents_count_per_department


def search_city(name: str, departement_name: str) -> dict[str, Any]:
    json_content = requests_get_meteo(
        "https://geocoding-api.open-meteo.com/v1/search",
        params={"name": name, "count": 20, "language": "fr", "format": "json"},
    )

    if "results" not in json_content:
        tokens = name.split("-")
        if len(tokens) == 1:
            error = f"No result found for '{name}'"
            raise CityError(error)
        # Try to remove the city suffix as a last resort
        return search_city("-".join(tokens[:1]), departement_name)

    results = [
        entry
        for entry in json_content["results"]
        if "country_code" in entry
        and (
            (entry["country_code"] == "GP" and departement_name == "Guadeloupe")
            or (entry["country_code"] == "MQ" and departement_name == "Martinique")
            or (entry["country_code"] == "GF" and departement_name == "Guyane")
            or (entry["country_code"] == "RE" and departement_name == "Réunion")
            or (
                entry["country_code"] == "PM"
                and departement_name == "Saint Pierre et Miquelon"
            )
            or (entry["country_code"] == "YT" and departement_name == "Mayotte")
            or (
                entry["country_code"] == "FR"
                and "admin2" in entry
                and entry["admin2"].lower() == departement_name.lower()
            )
        )
        and entry["feature_code"].startswith("PPL")
        and "elevation" in entry
    ]

    # Prioritize results which are exact matches
    if len(results) > 1:
        results = [r for r in results if r["name"].lower() == name.lower()]

    # Prioritize results which have the same admin4 name
    if len(results) > 1:
        results = [
            r for r in results if "admin4" in r and r["admin4"].lower() == name.lower()
        ]

    # Prioritize results which have a population count
    if len(results) > 1:
        results = [r for r in results if "population" in r]

    if len(results) > 1:
        error = (
            f"{len(results)} results were found for '{name}', "
            "don't know which one to pick"
        )
        raise CityError(error)

    if len(results) == 0:
        error = f"No result found for '{name}' after filter"
        raise CityError(error)

    return results[0]


def pick_cities(
    department_name: str,
    cities_json: list[Any],
    max_elevation: int,
    sample_size: int,
) -> list[City]:
    least_crowded_cities = []
    most_crowded_cities = []

    sorted_cities = sorted(cities_json, key=lambda c: c["population"])

    i = 0
    for i in range(len(sorted_cities)):
        city_json = sorted_cities[i]
        try:
            search_result = search_city(
                city_json["nom"],
                department_name,
            )
        except CityError as ce:
            if "No result found" in str(ce):
                continue
            raise
        city = City(
            search_result["name"],
            search_result["latitude"],
            search_result["longitude"],
            search_result["elevation"],
            city_json["population"],
            department_name,
        )

        if city.elevation < max_elevation:
            least_crowded_cities.append(city)

            if len(least_crowded_cities) == sample_size // 2:
                break

    for city_json in reversed(sorted_cities[i + 1 :]):
        try:
            search_result = search_city(
                city_json["nom"],
                department_name,
            )
        except CityError as ce:
            if "No result found" in str(ce):
                continue
            raise

        city = City(
            search_result["name"],
            search_result["latitude"],
            search_result["longitude"],
            search_result["elevation"],
            city_json["population"],
            department_name,
        )

        if city.elevation < max_elevation:
            most_crowded_cities.append(city)

            if len(least_crowded_cities) + len(most_crowded_cities) == sample_size:
                break

    picked_cities = least_crowded_cities + most_crowded_cities

    # Paris is an exception because it does not belong to any department,
    # it is considered to be its own department.
    # The other exception is Saint-Pierre et Miquelon which only has 2 cities
    if (
        len(picked_cities) < sample_size
        and department_name != "Paris"
        and len(sorted_cities) >= sample_size
    ):
        print(
            f"WARNING: didn't find enough cities for {department_name},",
            "will retry with an increased max elevation...",
        )
        return pick_cities(
            department_name,
            cities_json,
            max_elevation + 10,  # increase max elevation by 10 meters
            sample_size,
        )

    return picked_cities


def pick_cities_per_department(
    max_elevation: int,
    sample_size_per_department: int = 10,
) -> dict[str, list[City]]:
    cities_samples = {}
    for department_code, department_name in DEPARTMENTS.items():
        print(f"Processing {department_name} (code: {department_code})...")
        cities = requests.get(
            f"https://geo.api.gouv.fr/departements/{department_code}/communes",
            timeout=HTTP_TIMEOUT,
        )

        cities.raise_for_status()
        cities_json = cities.json()
        cities_json = [c for c in cities_json if "population" in c]
        picked_cities = pick_cities(
            department_name,
            cities_json,
            max_elevation,
            sample_size_per_department,
        )
        cities_samples[department_code] = picked_cities

    return cities_samples


def build_plot(
    output_file: Path,
    title: str,
    values_per_department: dict[str, Any],
    categories: list[tuple[str, str]],
    style: Style | None = None,
) -> None:
    # Exclude DOM-TOMs from the Jenks classification
    # because there are way too different.
    # They will be added back into the classifier but
    # won't influence the creation of clusters.
    values_per_department_without_dom_toms = {
        k: v for k, v in values_per_department.items() if len(k) < 3
    }
    classes = jenks_breaks(
        list(values_per_department_without_dom_toms.values()),
        n_classes=len(categories),
    )
    clusters = [{} for _ in range(len(categories))]

    for k, v in values_per_department.items():
        for i in range(len(categories)):
            if v <= classes[i + 1]:
                clusters[i][k] = v
                break

    values_per_dom_toms = {
        k: v for k, v in values_per_department.items() if len(k) == 3
    }

    for k, v in values_per_dom_toms.items():
        for i in range(len(categories)):
            if v <= classes[i + 1]:
                clusters[i][k] = v
                break
        else:
            clusters[-1][k] = v

    colors = [t[1] for t in categories]

    with NamedTemporaryFile() as css_file:
        config = Config()
        css_file.write(
            b".value { fill: black !important; } "
            b".departement { fill-opacity: 1 !important; }"
            b".tooltip-overlay { transform: translate(142px, 0px) !important; }",
        )

        for i, color in enumerate(colors):
            css_file.write(
                f".hatch-{i} {{ fill: url(#diagonalHatch{i}) !important; }}".encode(),
            )

            # This creates SVG patterns that will allow us to hatch
            # departments using JavaScript.
            config.defs.append(f"""
                <pattern id="diagonalHatch{i}"
                         width="10" height="10"
                         patternTransform="rotate(45 0 0)"
                         patternUnits="userSpaceOnUse">
                <rect x="0" y="0" width="10" height="10" style="fill: {color}"/>
                <line x1="0" y1="0"
                      x2="0" y2="10"
                      style="stroke:black; stroke-width:3" />
                </pattern>
            """)

        css_file.flush()
        config.css.append("file://" + css_file.name)  # type: ignore []

        france_map = FrenchMapDepartments(
            config=config,
            style=style
            or Style(background="#fcfcfc", plot_background="#fcfcfc", colors=colors),
        )
        france_map.title = title

        for i, cluster in enumerate(clusters):
            france_map.add(
                categories[i][0],
                [{"value": (k, v), "color": colors[i]} for k, v in cluster.items()],
            )

        with output_file.open("w") as map_file:
            map_file.write(france_map.render().decode())


def main() -> None:
    cities_per_department = pick_cities_per_department(
        max_elevation=400,
        sample_size_per_department=10,
    )

    average_winter_weather_per_city: dict[City, Weather] = {}
    average_summer_weather_per_city: dict[City, Weather] = {}

    for cities in cities_per_department.values():
        for city in cities:
            print(f"Processing weather of {city.name} in {city.departement}")
            weather = get_weather(city, date(1994, 1, 1), date(2023, 12, 31))
            city_winter_mean, city_summer_mean = compute_average_season_weather(weather)
            average_winter_weather_per_city[city] = city_winter_mean
            average_summer_weather_per_city[city] = city_summer_mean

    average_winter_weather_per_department: dict[str, Weather] = {}
    average_summer_weather_per_department: dict[str, Weather] = {}

    for department, cities in cities_per_department.items():
        all_cities_winter_weather = []
        all_cities_summer_weather = []

        for city in cities:
            all_cities_winter_weather.append(average_winter_weather_per_city[city])
            all_cities_summer_weather.append(average_summer_weather_per_city[city])

        average_department_winter_weather = compute_average_weather(
            all_cities_winter_weather,
        )
        average_department_summer_weather = compute_average_weather(
            all_cities_summer_weather,
        )
        average_winter_weather_per_department[department] = (
            average_department_winter_weather
        )
        average_summer_weather_per_department[department] = (
            average_department_summer_weather
        )

    mean_air_quality_per_department: dict[str, float] = {}
    max_air_quality_per_department: dict[str, float] = {}

    for department_code, cities in cities_per_department.items():
        department_air_quality: list[AirQuality] = []
        for city in cities:
            print(f"Processing air quality of {city.name} in {city.departement}")
            air_quality_mean = get_air_quality_mean(
                city,
                date(2022, 7, 29),
                date(2024, 7, 7),
            )
            department_air_quality.append(air_quality_mean)

        mean_air_quality_per_department[department_code] = mean(
            [aq.european_aqi_mean for aq in department_air_quality],
        )
        max_air_quality_per_department[department_code] = mean(
            [aq.european_aqi_max for aq in department_air_quality],
        )

    print("Processing soil pollution incidents count per department...")
    soil_pollution_incidents_count_per_department: dict[str, int] = (
        get_soil_pollution_incidents_count()
    )

    print("Processing natural disasters count per department...")
    natural_disasters_count_per_department: dict[str, int] = (
        get_referenced_natural_disaster_count()
    )

    print("Building maps...")
    build_plot(
        THIS_SCRIPT_LOCATION.joinpath("../_static/images/carte_qualite_de_lair.svg"),
        "Indice moyen de qualité de l'air (2022 - 2024)",
        mean_air_quality_per_department,
        [
            ("Très bon", "#43A047"),
            ("Bon", "#C0CA33"),
            ("Moyen", "#FDD835"),
            ("Mauvais", "#FB8C00"),
            ("Très mauvais", "#E53935"),
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_pics_de_pollution_de_lair.svg",
        ),
        "Pics de pollution de l'air (2022 - 2024)",
        max_air_quality_per_department,
        [
            ("Très bon", "#43A047"),
            ("Bon", "#C0CA33"),
            ("Moyen", "#FDD835"),
            ("Mauvais", "#FB8C00"),
            ("Très mauvais", "#E53935"),
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_ratio_precipitations_evapotranspiration.svg",
        ),
        "Ratio précipitations / évapotranspiration en été (1994 - 2023)",
        {
            k: v.rainfall_sum / v.et0_fao_evapotranspiration
            for k, v in average_summer_weather_per_department.items()
        },
        [
            ("Très mauvais", "#E53935"),
            ("Mauvais", "#FB8C00"),
            ("Moyen", "#FDD835"),
            ("Bon", "#C0CA33"),
            ("Très bon", "#43A047"),
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_temperature_ressentie_moyenne_ete.svg",
        ),
        "Température ressentie moyenne en été (1994 - 2023)",
        {
            k: v.apparent_temperature_mean
            for k, v in average_summer_weather_per_department.items()
        },
        [
            ("Le plus froid", "#FFCDD2"),  # 100
            ("Plutôt froid", "#EF9A9A"),  # 200
            ("Tempéré", "#EF5350"),  # 400
            ("Plutôt chaud", "#E53935"),  # 600
            ("Le plus chaud", "#C62828"),  # 800
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_temperature_ressentie_moyenne_hiver.svg",
        ),
        "Température ressentie moyenne en hiver (1994 - 2023)",
        {
            k: v.apparent_temperature_mean
            for k, v in average_winter_weather_per_department.items()
        },
        [
            ("Le plus froid", "#1565C0"),  # 800
            ("Plutôt froid", "#1E88E5"),  # 600
            ("Tempéré", "#42A5F5"),  # 400
            ("Plutôt chaud", "#90CAF9"),  # 200
            ("Le plus chaud", "#BBDEFB"),  # 100
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_temperature_ressentie_max_moyenne_ete.svg",
        ),
        "Température ressentie maximale journalière moyenne en été (1994 - 2023)",
        {
            k: v.apparent_temperature_max
            for k, v in average_summer_weather_per_department.items()
        },
        [
            ("Le plus froid", "#FFCDD2"),  # 100
            ("Plutôt froid", "#EF9A9A"),  # 200
            ("Tempéré", "#EF5350"),  # 400
            ("Plutôt chaud", "#E53935"),  # 600
            ("Le plus chaud", "#C62828"),  # 800
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_temperature_ressentie_min_moyenne_hiver.svg",
        ),
        "Température ressentie minimale journalière moyenne en hiver (1994 - 2023)",
        {
            k: v.apparent_temperature_min
            for k, v in average_winter_weather_per_department.items()
        },
        [
            ("Le plus froid", "#1565C0"),  # 800
            ("Plutôt froid", "#1E88E5"),  # 600
            ("Tempéré", "#42A5F5"),  # 400
            ("Plutôt chaud", "#90CAF9"),  # 200
            ("Le plus chaud", "#BBDEFB"),  # 100
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_ensolleillement_moyen_hiver.svg",
        ),
        "Ensolleillement moyen en heures en hiver (1994 - 2023)",
        {
            k: v.sunshine_duration
            for k, v in average_winter_weather_per_department.items()
        },
        [
            ("Très faible", "#FFF9C4"),  # 100
            ("Faible", "#FFF59D"),  # 200
            ("Moyen", "#FFEE58"),  # 400
            ("Important", "#FDD835"),  # 600
            ("Très important", "#F9A825"),  # 800
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_ensolleillement_moyen_ete.svg",
        ),
        "Ensolleillement moyen en heures en été (1994 - 2023)",
        {
            k: v.sunshine_duration
            for k, v in average_summer_weather_per_department.items()
        },
        [
            ("Très faible", "#FFF9C4"),  # 100
            ("Faible", "#FFF59D"),  # 200
            ("Moyen", "#FFEE58"),  # 400
            ("Important", "#FDD835"),  # 600
            ("Très important", "#F9A825"),  # 800
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_duree_moyenne_journee_hiver.svg",
        ),
        "Durée moyenne d'une journée hivernale en heures (1994 - 2023)",
        {
            k: v.daylight_duration
            for k, v in average_winter_weather_per_department.items()
        },
        [
            ("Très faible", "#FFF9C4"),  # 100
            ("Faible", "#FFF59D"),  # 200
            ("Moyen", "#FFEE58"),  # 400
            ("Important", "#FDD835"),  # 600
            ("Très important", "#F9A825"),  # 800
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_duree_moyenne_journee_ete.svg",
        ),
        "Durée moyenne d'une journée estivale en heures (1994 - 2023)",
        {
            k: v.daylight_duration
            for k, v in average_summer_weather_per_department.items()
        },
        [
            ("Très faible", "#FFF9C4"),  # 100
            ("Faible", "#FFF59D"),  # 200
            ("Moyen", "#FFEE58"),  # 400
            ("Important", "#FDD835"),  # 600
            ("Très important", "#F9A825"),  # 800
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_vitesse_du_vent_ete.svg",
        ),
        "Vitesse moyenne du vent 10m au dessus du sol en km/h en été (1994 - 2023)",
        {
            k: v.wind_speed_10m_max
            for k, v in average_summer_weather_per_department.items()
        },
        [
            ("Très faible", "#B2DFDB"),  # 100
            ("Faible", "#80CBC4"),  # 200
            ("Moyen", "#26A69A"),  # 400
            ("Important", "#00897B"),  # 600
            ("Très important", "#00695C"),  # 800
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_vitesse_du_vent_hiver.svg",
        ),
        "Vitesse moyenne du vent 10m au dessus du sol en km/h en hiver (1994 - 2023)",
        {
            k: v.wind_speed_10m_max
            for k, v in average_winter_weather_per_department.items()
        },
        [
            ("Très faible", "#B2DFDB"),  # 100
            ("Faible", "#80CBC4"),  # 200
            ("Moyen", "#26A69A"),  # 400
            ("Important", "#00897B"),  # 600
            ("Très important", "#00695C"),  # 800
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_chutes_de_neige_hiver.svg",
        ),
        "Chutes de neige moyennes sur toute la période hivernale en mm (1994 - 2023)",
        {k: v.snowfall_sum for k, v in average_winter_weather_per_department.items()},
        [
            ("Très faibles", "#B2EBF2"),  # 100
            ("Faibles", "#80DEEA"),  # 200
            ("Moyennes", "#26C6DA"),  # 400
            ("Importantes", "#00ACC1"),  # 600
            ("Très importantes", "#00838F"),  # 800
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath("../_static/images/carte_pollution_des_sols.svg"),
        "Nombre d'incidents recensés dans la base BASOL ayants entrainé une pollution "
        "des sols",
        soil_pollution_incidents_count_per_department,
        [
            ("Très faible", "#43A047"),
            ("Faible", "#C0CA33"),
            ("Moyen", "#FDD835"),
            ("Important", "#FB8C00"),
            ("Très important", "#E53935"),
        ],
    )

    build_plot(
        THIS_SCRIPT_LOCATION.joinpath(
            "../_static/images/carte_catastrophes_naturelles.svg",
        ),
        "Nombre de catastrophes naturelles recensées dans la base GASPAR",
        natural_disasters_count_per_department,
        [
            ("Très faible", "#43A047"),
            ("Faible", "#C0CA33"),
            ("Moyen", "#FDD835"),
            ("Important", "#FB8C00"),
            ("Très important", "#E53935"),
        ],
    )


if __name__ == "__main__":
    main()
